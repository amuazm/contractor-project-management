import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook

df = pd.read_excel("./Files/Output/Result_ARIMA.xlsx", sheet_name="Reports")

# Keep actual data, remove forecasted
df = df[df["Remark"] == "Actual Date"]
# Keep latest record for each project
df = df.drop_duplicates(subset=["Project ID"], keep="last")
# Keep Project ID, VAC, VAC(t)
df = df[["Project ID", "VAC", "VAC(t)"]]

# Normalisation
scaler = MinMaxScaler()
scaler.fit(df[["VAC", "VAC(t)"]])
df[["VAC(n)", "VAC(t)(n)"]] = scaler.transform(df[["VAC", "VAC(t)"]])

km = KMeans(n_clusters=3)
y_predicted = km.fit_predict(df[["VAC(n)", "VAC(t)(n)"]])

d = {}
for i in range(len(km.cluster_centers_)):
    d[i] = [np.linalg.norm(km.cluster_centers_[i] - [0, 0])]
center_coords = []
for i in d.values():
    center_coords.append(i[0])
center_coords = sorted(center_coords)
for i in d:
    d[i].append(center_coords.index(d[i]))
labels = {0: "Higher Risk", 1: "Intermediate Risk", 2: "Lower Risk"}
for i in d:
    d[i][1] = labels[d[i][1]]
y_predicted = list(y_predicted)
for i in range(len(y_predicted)):
    y_predicted[i] = d[y_predicted[i]][1]

df["Risk"] = y_predicted

print(df)

wb_result = load_workbook("./Files/Output/Result.xlsx")

wb_result.create_sheet("Clustering")
ws_clustering = wb_result["Clustering"]

ws_clustering.append(list(df))
for index, row in df.iterrows():
    l = []
    for column in list(df):
        l.append(row[column])
    ws_clustering.append(l)

wb_result.save("./Files/Output/Result_Clustering.xlsx")

# df1 = df[df["Risk"] == "Higher Risk"]
# df2 = df[df["Risk"] == "Intermediate Risk"]
# df3 = df[df["Risk"] == "Lower Risk"]
# plt.scatter(df1["VAC(n)"], df1["VAC(t)(n)"], label="Higher Risk", color="red")
# plt.scatter(df2["VAC(n)"], df2["VAC(t)(n)"], label="Intermediate Risk", color="yellow")
# plt.scatter(df3["VAC(n)"], df3["VAC(t)(n)"], label="Lower Risk", color="green")
# plt.xlabel("VAC(n)")
# plt.ylabel("VAC(t)(n)")
# plt.legend()
# plt.show()