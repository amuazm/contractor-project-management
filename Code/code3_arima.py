# ARIMA

import shutil
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from statsmodels.tsa.arima.model import ARIMA
from dateutil import relativedelta
from openpyxl import load_workbook

# Copy file
src = "./Excel Files/Output/Result_EAC.xlsx"
dest = "./Excel Files/Output/Result_ARIMA.xlsx"
shutil.copyfile(src, dest)

# Load Result.xlsx
wb_result = load_workbook(dest)
ws_budget = wb_result["Budget"]
ws_reports = wb_result["Reports"]

# Add "Remark" column
ws_reports.insert_cols(16)
ws_reports["P1"] = "Remark"
for row in ws_reports.iter_rows(min_row=2):
    row[15].value = "Actual Date"

# Get Overall Budgets
overall_budgets = {}
for row in ws_budget.iter_rows(min_row=2):
    overall_budgets[row[0].value] = row[4].value
# Get Durations
project_durations = {}
for row in ws_budget.iter_rows(min_row=2):
    project_durations[row[0].value] = row[2].value
# Get Monthly Budgets
monthly_budgets = {}
for row in ws_budget.iter_rows(min_row=2):
    monthly_budgets[row[0].value] = row[3].value

# Get Reports sheet as dataframe for input into ARIMA function
df_reports = pd.read_excel(dest, sheet_name="Reports")

# Perform ARIMA for each project
for project_id in overall_budgets:
    # # Test only one Project
    # if project_id != "Project 269":
    #     continue
    print("\n\n\n\n\n\n==========================", project_id, "==========================")
    # Filter project ID
    df = df_reports[df_reports["Project ID"] == project_id]
    # Filter Date (Time-Series), ACWP, and BCWP (forecasting these 2 metrics)
    df = df[["Date", "ACWP", "BCWP"]]
    df = df.set_index(["Date"])
    # Floor date to start of month for frequency
    df.index = pd.to_datetime(df.index) - pd.tseries.offsets.MonthBegin(1)
    df = df.asfreq(pd.infer_freq(df.index))

    # # Check ACWP and BCWP df before sending into ARIMA
    # print(df[["ACWP"]])
    # print(df[["BCWP"]])
    # plt.plot(df.index, df["ACWP"])
    # plt.plot(df.index, df["BCWP"])
    # plt.show()

    # Calculate number of months to predict based off VAC(t)
    months_passed = len(df)
    project_duration = project_durations[project_id]
    planned_months_left = project_duration - months_passed
    estimated_month_variance = df_reports.loc[df_reports["Project ID"] == project_id]["VAC(t)"].iloc[-1] # Negative means extra months estimated (behind schedule). This line of code just gets the latest VAC(t) for the project
    months_to_predict = np.ceil(project_duration - months_passed - estimated_month_variance) # basically Duration (Months) - Months Passed - VAC(t). Another way is to just get EAC(t) - Months Passed, I think, which I didn't consider

    # Start and ending dates of prediction
    pred_start_date = df.index[-1:]
    pred_start_date = pred_start_date.to_pydatetime()[0] + relativedelta.relativedelta(months=1)
    pred_end_date = pred_start_date + relativedelta.relativedelta(months=months_to_predict)


    # Split ACWP and BCWP
    l = []
    l.append(df[["ACWP"]])
    l.append(df[["BCWP"]])

    l_predictions = []
    try:
        for df_to_predict in l:
            # # Maximum number of lags possible
            # lags = len(df.index)//3
            # # Check ACF and PACF
            # acf_plot = plot_acf(df2, lags=lags)
            # pacf_plot = plot_pacf(df2, lags=lags, method="ywm")
            # plt.show()

            model = ARIMA(df_to_predict, order=(1, 2, 0))
            model_fit = model.fit()

            predictions = model_fit.predict(start=pred_start_date, end=pred_end_date)
            predictions = predictions.to_frame()
            predictions.index.name = "Date"
            predictions.columns = [list(df_to_predict)[0]]

            l_predictions.append(predictions)

        # combined_for_testing = pd.concat([df, pd.concat([l2[0], l2[1]], axis=1)])
        # plt.plot(combined_for_testing[["ACWP"]], label="ACWP")
        # plt.plot(combined_for_testing[["BCWP"]], label="BCWP")
        # plt.legend()
        # plt.show()
        # plt.clf()

        df = pd.concat([l_predictions[0], l_predictions[1]], axis=1)

        # Place into excel file
        reached_one_hundred_percent = False
        prev_acwp = 0
        prev_bcwp = 0
        for index, row in df.iterrows():
            acwp = row["ACWP"]
            bcwp = row["BCWP"]
            if reached_one_hundred_percent == False:
                # Never allow acwp or bcwp to have a negative slope
                if acwp < prev_acwp:
                    acwp = prev_acwp
                if bcwp < prev_bcwp:
                    bcwp = prev_bcwp

                # Mark to stop if project is complete (bcwp = overall budget)
                if bcwp >= overall_budgets[project_id]:
                    bcwp = overall_budgets[project_id]
                    reached_one_hundred_percent = True

                ws_reports.append([project_id, index.date(), "", "", acwp, bcwp, "", "", "", "", "", "", "", "", "", "Forecasted"])
                prev_acwp = acwp
                prev_bcwp = bcwp
    except Exception as e:
        print(e)

months_passed = {}
for row in ws_reports.iter_rows(min_row=2):
    project_id = row[0].value
    # Months Passed Calculation
    if project_id not in months_passed:
        months_passed[project_id] = 1
    else:
        months_passed[project_id] += 1

    if row[15].value == "Forecasted":
        # Months Passed
        row[2].value = months_passed[project_id]

        # Completion
        row[3].value = row[5].value / overall_budgets[project_id]

        # BCWS
        bcws = monthly_budgets[project_id] * months_passed[project_id]
        if bcws >= overall_budgets[project_id]:
            bcws = overall_budgets[project_id]
        row[6].value = bcws

        # CPI
        acwp = row[4].value
        bcwp = row[5].value
        row[7].value = bcwp / acwp
        cpi = row[7].value

        # CV
        row[8].value = bcwp - acwp

        # SPI
        row[9].value = bcwp / bcws
        spi = row[9].value

        # SV
        row[10].value = bcwp - bcws

        # EAC
        row[11].value = acwp + (overall_budgets[project_id] - bcwp) / cpi
        eac = row[11].value

        # EAC(t)
        row[12].value = months_passed[project_id] + (max(project_durations[project_id], months_passed[project_id]) - months_passed[project_id] * spi) / spi
        eac_t = row[12].value

        # VAC
        row[13].value = overall_budgets[project_id] - eac

        # VAC(t)
        row[14].value = project_durations[project_id] - eac_t

        row[3].style = "Percent"
        row[4].style = "Currency"
        row[5].style = "Currency"
        row[6].style = "Currency"
        row[8].style = "Currency"
        row[10].style = "Currency"
        row[11].style = "Currency"
        row[13].style = "Currency"

# Save as Result_ARIMA.xlsx
wb_result.save(dest)