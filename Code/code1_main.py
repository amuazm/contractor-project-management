import shutil
from openpyxl import Workbook, load_workbook
import datetime
import numpy as np
from dateutil.relativedelta import relativedelta
import names
# import matplotlib.pyplot as plt

# Copy dataset to refer from as backup
src = "./Excel Files/Input/Budget Dataset Modified.xlsx"
dest = "./Excel Files/Output/Data.xlsx"
shutil.copyfile(src, dest)

# Create new empty workbook to start from
wb_result = Workbook()
# Sheets
wb_result.create_sheet("Project Details")
wb_result.create_sheet("Budget")
wb_result.create_sheet("Categories")
wb_result.create_sheet("Cash Outflow")
wb_result.create_sheet("Cash Inflow")
wb_result.create_sheet("Reports")
ws_project_details = wb_result["Project Details"]
ws_budget = wb_result["Budget"]
ws_categories = wb_result["Categories"]
ws_cash_outflow = wb_result["Cash Outflow"]
ws_cash_inflow = wb_result["Cash Inflow"]
ws_reports = wb_result["Reports"]
# Delete default sheet
wb_result.remove(wb_result["Sheet"])

# Project Details
if True:
    # Load dataset to refer from
    wb_data = load_workbook(dest)
    # Copy entire Project Details sheet row by row from "data" to "result"
    for row in wb_data["Project Details"].iter_rows():
        l = []
        for cell in row:
            l.append(cell.value)
        ws_project_details.append(l)
    wb_result.close()
    # Styles
    for row in ws_project_details.iter_rows(min_row=2):
        row[11].style = "Currency"

# Project ID list
project_ids = ws_project_details["A"][1:]
project_ids = [i.value for i in project_ids]
# Overall budgets dict
overall_budgets = {}
for row in ws_project_details.iter_rows(min_row=2):
    overall_budgets[row[0].value] = row[11].value

# Budget
if True:
    # Headers
    ws_budget.append(["Project ID", "Start Date", "Duration (Months)", "Budgeted Cost per Month", "Overall Budget", "Projected Income"])

    i = 0
    for row in ws_budget.iter_rows(min_row=2, max_row=len(project_ids) + 1, max_col=6):
        # Project IDs
        row[0].value = project_ids[i]
        i += 1

        # Start Date
        # Generate a random date from 2 years ago to 6 months ago (for 6-24 months of data for each project)
        start = datetime.date.today() - relativedelta(years=2)
        end = datetime.date.today() - relativedelta(months=6)
        delta = end - start
        r = np.random.randint(0, delta.days + 1)
        new = start + datetime.timedelta(days=r)
        row[1].value = new

        # Duration (Months)
        # 2.5 to 5 years
        row[2].value = np.random.randint(30, 60)

        # Cost per Month
        row[3].value = overall_budgets[row[0].value] / row[2].value
        row[3].style = "Currency"

        # Overall Budget
        row[4].value = overall_budgets[row[0].value]
        row[4].style = "Currency"

        # Contract Pay
        row[5].value = overall_budgets[row[0].value] * np.random.randint(1020, 1080)/1000
        row[5].style = "Currency"

# Categories
if True:
    # Headers
    ws_categories.append(["Project ID", "Category", "Category Budget"])

    categories = [
        'Direct Labour',
        'Supplied Labour',
        'Sub-contractor',
        'Other Materials',
        'Small Tools & Safety Item',
        'Other Consumable',
        'Transportation',
        'Repair & Maintenance',
        'Site Office Expense',
        'Food, Refreshment & Entertainment',
        'Travelling & Vehicles',
        'Main Steel Materials',
        'Stainless Steel Materials',
        'Aluminium Materials',
        'Equipment',
        'Supervision',
        'Insurance'
    ]

    # Randomly spread out percentages of how much each category is budgeted. Maximum of 100%. Percentages can be found in k_nums.
    for id in project_ids:
        n, k = overall_budgets[id], len(categories)
        vals = np.random.default_rng().dirichlet(np.ones(k), size=1)
        k_nums = [round(v) for v in vals[0]*n]
        i = 0
        for category in categories:
            ws_categories.append([id, category, k_nums[i]])
            i += 1

    # Styles
    for row in ws_categories.iter_rows(min_row=2):
        row[2].style = "Currency"

# Start dates dict
start_dates = {}
for row in ws_budget.iter_rows(min_row=2):
    start_dates[row[0].value] = row[1].value
# Project durations dict
project_durations = {}
for row in ws_budget.iter_rows(min_row=2):
    project_durations[row[0].value] = row[2].value
# Budgets by category nested dict
budgets_by_cat = {}
for row in ws_categories.iter_rows(min_row=2):
    if row[0].value not in budgets_by_cat:
        budgets_by_cat[row[0].value] = {}
    budgets_by_cat[row[0].value][row[1].value] = row[2].value

# Cash Outflow
if True:
    # Headers
    ws_cash_outflow.append(["Project ID", "Date", "Category", "Actual Category Monthly Cost"])

    for id in project_ids:
        cash_outflow_date = start_dates[id]
        # First cash outflow date will be 1 month after project start date
        cash_outflow_date += relativedelta(months=1)
        # Add 1 month to date until today's date
        # Random Project Cost
        r1 = np.random.randint(700, 1300 + 1) / 1000
        while(datetime.date.today() - cash_outflow_date > datetime.timedelta(days=0)):
            # Random Date Cost
            r2 = np.random.randint(800, 1200 + 1) / 1000
            for category in categories:
                # Random Category Cost
                r3 = np.random.randint(900, 1100 + 1) / 1000
                # Actual Category Monthly Cost = Category Budget / Project Duration
                monthly_cash_outflow = budgets_by_cat[id][category] / project_durations[id] * r1 * r2 * r3
                ws_cash_outflow.append([id, cash_outflow_date, category, monthly_cash_outflow])
            cash_outflow_date += relativedelta(months=1)
    # Styles
    for row in ws_cash_outflow.iter_rows(min_row=2):
        row[3].style = "Currency"

# Actual costs by date and category nested dict
actual_by_cat_date = {}
for row in ws_cash_outflow.iter_rows(min_row=2):
    if row[0].value not in actual_by_cat_date:
        actual_by_cat_date[row[0].value] = {}
    if row[1].value not in actual_by_cat_date[row[0].value]:
        actual_by_cat_date[row[0].value][row[1].value] = {}
    actual_by_cat_date[row[0].value][row[1].value][row[2].value] = row[3].value
# Monthly budgets dict
monthly_budgets = {}
for row in ws_budget.iter_rows(min_row=2):
    monthly_budgets[row[0].value] = row[3].value

# Reports
if True:
    # Headers
    ws_reports.append(["Project ID", "Date", "Completion", "ACWP", "BCWP", "BCWS", "CPI", "CV", "SPI", "SV"])

    # Project ID
    for id in actual_by_cat_date:
        months_passed = 1
        acwp = 0
        # Random Project Completion
        r1 = np.random.randint(950, 1050 + 1) / 1000
        r2 = r1
        completion = 0
        # Date of Report
        for date in actual_by_cat_date[id]:
            # Completion
            if (completion > 1):
                break
            completion = months_passed * (1/project_durations[id]) * r1
            if (completion > 1):
                completion = 1
            r1 *= r2
            if r2 < 1:
                r2 *= 1.001
            else:
                r2 *= 0.999

            # ACWP
            date_cost = 0
            for category in actual_by_cat_date[id][date]:
                    date_cost += actual_by_cat_date[id][date][category]
            acwp += date_cost

            # BCWP
            bcwp = overall_budgets[id] * completion

            # BCWS
            bcws = months_passed * monthly_budgets[id]

            # CPI
            cpi = bcwp / acwp

            # CV
            cv = bcwp - acwp

            # SPI
            spi = bcwp / bcws

            # SV
            sv = bcwp - bcws

            ws_reports.append([id, date, completion, acwp, bcwp, bcws, cpi, cv, spi, sv])
            months_passed += 1

    # Styles
    for row in ws_reports.iter_rows(min_row=2):
        row[2].style = "Percent"
        row[3].style = "Currency"
        row[4].style = "Currency"
        row[5].style = "Currency"
        row[7].style = "Currency"
        row[9].style = "Currency"

# Project completions nested dict
project_completions = {}
for row in ws_reports.iter_rows(min_row=2):
    if row[0].value not in project_completions:
        project_completions[row[0].value] = {}
    project_completions[row[0].value][row[1].value] = row[2].value
# Projected incomes dict
projected_incomes = {}
for row in ws_budget.iter_rows(min_row=2):
    projected_incomes[row[0].value] = row[5].value

# Cash Inflow
if True:
    # Headers
    ws_cash_inflow.append(["Project ID", "Date", "Income"])

    for id in project_completions:
        # For each project, Income is separated into 5 sections
        # 20% of Projected Income at 0%, 25%, 50%, 75%, and 100% completion each. 100% shouldnt be possible here but added just in case.
        income = projected_incomes[id] / 5
        milestones = []
        for date in project_completions[id]:
            completion = project_completions[id][date]

            append = True

            if milestones == []:
                pass
            elif completion >= 0.25 and True not in [i >= 0.25 for i in milestones]:
                pass
            elif completion >= 0.5 and True not in [i >= 0.5 for i in milestones]:
                pass
            elif completion >= 0.75 and True not in [i >= 0.75 for i in milestones]:
                pass
            elif completion >= 1 and True not in [i >= 1 for i in milestones]:
                pass
            else:
                append = False

            if append:
                milestones.append(completion)
                ws_cash_inflow.append([id, date, income])
    
    # Styles
    for row in ws_cash_inflow.iter_rows(min_row=2):
        row[2].style = "Currency"

# Project completions latest dict
project_completions_latest = {}
for row in ws_reports.iter_rows(min_row=2):
    project_completions_latest[row[0].value] = row[2].value

# Other
if True:
    # Change Project Status
    # 0-15 = Planning, 15-30 = Planned, 30-70 = Execution, 70-100 = Finishing, 100 = Finished
    for row in ws_project_details.iter_rows(min_row=2):
        if project_completions_latest[row[0].value] >= 1:
            row[3].value = "Finished"
        elif project_completions_latest[row[0].value] >= 0.7:
            row[3].value = "Finishing"
        elif project_completions_latest[row[0].value] >= 0.3:
            row[3].value = "Execution"
        elif project_completions_latest[row[0].value] >= 0.15:
            row[3].value = "Planned"
        else:
            row[3].value = "Planning"

    # Names
    ws_project_details["G1"] = "Project Manager"
    r = 0
    for row in ws_project_details.iter_rows(min_row=2):
        if r == 0:
            r = np.random.randint(1, 6)
            name = names.get_full_name()
        row[6].value = name
        r -= 1

    # Regions
    # Hardcoded. Manually researched and typed.
    regions = [
    "North-East",
    "North-East",
    "Central",
    "West",
    "West",
    "West",
    "Central",
    "Central",
    "West",
    "Central",
    "North",
    "East",
    "West",
    "West",
    "West",
    "Central",
    "Central",
    "North-East",
    "West",
    "West",
    "Central",
    "North",
    "North",
    "Central",
    "Central",
    "Central",
    "Central",
    "North-East",
    "North-East",
    "Central",
    "Central",
    "Central",
    "East",
    "East",
    "East"
    ]
    ws_project_details.insert_cols(4, 1)
    ws_project_details["D1"] = "Region"
    i = 0
    for row in ws_project_details.iter_rows(min_row=2):
        row[3].value = regions[i]
        i += 1

# Save result
wb_result.save("./Excel Files/Output/Result.xlsx")